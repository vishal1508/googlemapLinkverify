from flask import Flask, render_template, request, send_file
import os
import re
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Configure upload and output directories
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

# Function to check if the URL is a valid Google Maps address
def is_google_maps_address(url):
    try:
        # Parse the URL
        parsed_url = urlparse(url)
        
        # Check if the domain is google.com and if the path contains 'maps'
        if 'google.com' not in parsed_url.netloc or 'maps' not in parsed_url.path:
            return False
        
        # Check if coordinates are included in the URL after '@'
        if '@' in parsed_url.path:
            coordinates = parsed_url.path.split('@')[1]  # Get the part after '@'
            
            # Now, split by comma and take only the first two elements (latitude and longitude)
            lat_lng = coordinates.split(',')[:2]  # Take only latitude and longitude
            
            if len(lat_lng) == 2 and re.match(r"^\-?\d+(\.\d+)?$", lat_lng[0]) and re.match(r"^\-?\d+(\.\d+)?$", lat_lng[1]):
                return True
        
        # Check for coordinates in the query parameter (q=latitude,longitude)
        if 'q=' in parsed_url.query:
            query_params = parsed_url.query.split('&')
            for param in query_params:
                if param.startswith('q='):
                    coordinates = param.split('=')[1]  # Extract the value after 'q='
                    lat_lng = coordinates.split(',')[:2]  # Get latitude and longitude

                    if len(lat_lng) == 2 and re.match(r"^\-?\d+(\.\d+)?$", lat_lng[0]) and re.match(r"^\-?\d+(\.\d+)?$", lat_lng[1]):
                        return True
        
        return False

    except Exception as e:
        # Handle any exceptions that might occur during URL parsing
        return False

# Function to extract the URL from a hyperlink (Excel formula or text)
def extract_url_from_hyperlink(cell):
    if cell.hyperlink:
        return cell.hyperlink.target
    return cell.value

@app.route('/')
def upload_file():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return "No file uploaded", 400

    file = request.files['file']
    if file.filename == '':
        return "No file selected", 400

    if file:
        # Save the uploaded file
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        # Process the Excel file
        output_filename = f"processed_{file.filename}"
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)

        # Read the Excel file using openpyxl to extract hyperlinks
        wb = load_workbook(file_path)
        sheet = wb.active

        # Assuming 'Reference Link' is the header, find the column index for 'Reference Link'
        columns = [cell.value for cell in sheet[1]]
        if 'Reference Link' not in columns:
            return "The Excel file must contain a column named 'Reference Link'", 400

        reference_link_col_index = columns.index('Reference Link') + 1  # Excel columns are 1-indexed

        # Add a new column header for 'Valid Map Address' if not already present
        if 'Valid Map Address' not in columns:
            new_col_index = len(columns) + 1
            sheet.cell(row=1, column=new_col_index, value='Valid Map Address')
        else:
            new_col_index = columns.index('Valid Map Address') + 1

        # Create a list of URLs and check their validity
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            # Get the URL from the 'Reference Link' column
            link_cell = row[reference_link_col_index - 1]  # Adjust for 0-indexed list
            raw_url = extract_url_from_hyperlink(link_cell)  # Extract raw URL

            # Check if the URL is a valid Google Maps address
            if raw_url:
                valid = is_google_maps_address(raw_url)
            else:
                valid = None  # Keep empty if no data in 'Reference Link'

            # Write the result in the 'Valid Map Address' column
            sheet.cell(row=row_idx, column=new_col_index, value=valid)

        # Save the processed file
        wb.save(output_path)

        # Send the processed file for download
        return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
